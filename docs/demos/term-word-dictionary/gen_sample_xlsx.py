# -*- coding: utf-8 -*-
"""Generate term_word import sample xlsx for review."""
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "term_word_import_sample.xlsx"

# 走LLM tip — 与前端表头 tip 文案保持一致
USE_LLM_TIP = (
    "是否需要走LLM判断，非直译，而是含有指代、有词性、需分场景判断等"
)

HEADERS = [
    "词片",
    "翻译",
    "翻译类型",
    "可见范围",
    "comment",
    "领域",
    "缩写",
    "走LLM",
    "使用场景与注意事项",
    "翻译状态",
]

# 走LLM 列用 Excel 布尔 TRUE/FALSE（对应源表「伪代码术语…」填「是」→ TRUE）
ROWS = [
    [
        "获取/检索",
        "v. retrieve\nn. retrieval",
        "英文",
        "",
        "",
        "",
        "",
        True,
        (
            "v.不要用obtain、get、got、parsing、fetching\n"
            "n.不要用acquisition\n"
            "“retrieve” 侧重于从特定存储位置取回；计算机领域常用。\n"
            "示例：\n"
            "xxx获取失败 → Failed to retrieve xxx\n"
            "xxx获取错误 → Error in retrieving xxx\n"
            "xxx获取不到 → xxx cannot be retrieved"
        ),
        "已审核",
    ],
    [
        "检索器",
        "Retriever",
        "英文",
        "",
        "",
        "数据库",
        "",
        False,
        "信息检索领域专指按索引/路径从存储层提取数据的机制。",
        "已审核",
    ],
    [
        "常驻检索器",
        "Resident retriever",
        "英文",
        "",
        "",
        "数据库",
        "",
        False,
        "不要用 Resident searcher。\nSearcher 更像搜索栏动作/组件；Retriever 强调按路径提取。",
        "已审核",
    ],
    [
        "画面",
        "graph",
        "英文",
        "",
        "",
        "图形",
        "",
        False,
        "不要用 screen、image（自动成图的图也指画面）。",
        "已审核",
    ],
    [
        "图元",
        "icon",
        "英文",
        "",
        "",
        "图形",
        "",
        False,
        "不要用 elements、graphic elements、entity、Metafile、graph icon、picture icons、Graphics 等。",
        "已审核",
    ],
]


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "术语字典导入"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei", size=11)
    cell_font = Font(name="Microsoft YaHei", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    alt = PatternFill("solid", fgColor="F6F8FA")

    use_llm_col = HEADERS.index("走LLM") + 1

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        if c == use_llm_col:
            cell.comment = Comment(USE_LLM_TIP, "术语字典", height=80, width=280)

    for r, row in enumerate(ROWS, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = cell_font
            cell.alignment = center if c == use_llm_col else wrap
            cell.border = thin
            if r % 2 == 0:
                cell.fill = alt

    widths = [14, 22, 12, 12, 12, 10, 10, 10, 48, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, 7):
        ws.row_dimensions[r].height = 72 if r == 2 else 48

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{1 + len(ROWS)}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("填写说明", 0)
    ws2["A1"] = "术语字典标准导入模板（样板）"
    ws2["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E79")
    notes = [
        "",
        "用途：人类维护词片真相源；导入后 Agent Grep 使用「词片→翻译」，拼装/翻译时读取「使用场景与注意事项」。",
        "正式导入只认本表「术语字典导入」sheet 的表头列名（勿改列名）。",
        "",
        "列说明：",
        "词片* — 中文词片原文（Grep 键）",
        "翻译* — 目标语译法；可含词性多行（如 v. / n.）",
        "翻译类型* — 与系统语种名一致，如 英文 / 俄文",
        "可见范围 — 部门可见范围，可空",
        "comment — 消歧键（与词条 comment 对齐，如 buttonName）；可空",
        "领域 — 如 数据库、图形（入库 category）",
        "缩写 — 英文缩写（入库 remark2）",
        f"走LLM — 布尔 TRUE/FALSE（亦接受 是/否、1/0）；{USE_LLM_TIP}",
        "         对应源表「伪代码术语(有指代、有词性、需分场景判断等)」；入库 use_llm",
        "         TRUE=该词片命中后仍应走 LLM 判断，不可当纯直译词表替换",
        "使用场景与注意事项 — 禁用译法、长度限制、示例句（入库 usage_notes；Agent 主读）",
        "翻译状态 — 未翻译/待审核/审核不通过/已审核 或 0/1/2/3；Agent 仅使用「已审核」",
        "",
        "样板 5 行取自《常用注意要点清单》「通用术语」sheet，状态均为已审核。",
        "其中「获取/检索」走LLM=TRUE（源表填「是」）；其余为 FALSE。",
        "原清单其它 sheet（流程草稿、模块对照等）不导入；可用产品内 notes-adapt 只转换「通用术语」。",
        "",
        "重复键策略：同（词片, comment, 翻译类型）已存在则跳过并记入导入报告。",
        "",
        "字段入库映射：词片→word，翻译→translate，翻译类型→target_lang，可见范围→department，",
        "comment→comment，领域→category，缩写→abbr，走LLM→use_llm（boolean），",
        "使用场景与注意事项→usage_notes，翻译状态→status。",
    ]
    for i, line in enumerate(notes, 2):
        ws2.cell(i, 1, line).font = Font(name="Microsoft YaHei", size=10)
    ws2.column_dimensions["A"].width = 108

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
