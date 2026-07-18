"""comment_rule Excel 导入解析。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.shared.comment_rule.format import infer_case_type, infer_prefer_abbr


def parse_comment_rule_workbook(data: bytes) -> list[dict[str, Any]]:
    """解析 comment/词条来源/场景/规则 表 → 行 dict 列表。"""
    wb = load_workbook(BytesIO(data), data_only=True)
    sheet_name = next(
        (n for n in wb.sheetnames if "comment" in str(n).lower()),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    header = [str(h or "").strip() for h in header_row]

    def idx(*names: str) -> int:
        for name in names:
            try:
                return header.index(name)
            except ValueError:
                continue
        return -1

    i_comment = idx("comment")
    i_source = idx("词条来源")
    i_scene = idx("场景")
    i_tips = idx("规则", "翻译补充要点")
    if i_comment < 0:
        raise ValueError("缺少表头列 comment")

    out: list[dict[str, Any]] = []
    for row in rows_iter:
        if not row:
            continue
        comment = str(row[i_comment] or "").strip() if i_comment >= 0 else ""
        if not comment:
            continue
        source = (
            str(row[i_source] or "").strip()
            if i_source >= 0 and i_source < len(row)
            else ""
        )
        scene = (
            str(row[i_scene] or "").strip()
            if i_scene >= 0 and i_scene < len(row)
            else ""
        )
        tips = (
            str(row[i_tips] or "").strip()
            if i_tips >= 0 and i_tips < len(row)
            else ""
        )
        out.append(
            {
                "comment_key": comment,
                "entry_source": source or None,
                "scene": scene or None,
                "rule_text": tips or None,
                "prefer_abbr": infer_prefer_abbr(scene, tips),
                "case_type": infer_case_type(scene, tips),
            }
        )
    return out
