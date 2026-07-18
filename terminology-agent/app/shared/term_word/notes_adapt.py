"""注意事项清单 → 标准术语字典行（只读「通用术语」）。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.shared.term_word.excel_columns import NOTES_HEADER_MAP, NOTES_SHEET_NAME
from app.shared.term_word.excel_io import parse_use_llm


def adapt_notes_workbook(
    content: bytes,
    *,
    target_lang: str = "英文",
    default_status: str = "1",
) -> list[dict[str, Any]]:
    """解析注意事项类 xlsx 的「通用术语」sheet → 标准行。

    Args:
        content: 上传文件 bytes。
        target_lang: 默认翻译类型（源表通常只有英文列）。
        default_status: 默认翻译状态（有翻译时建议待审核 1）。

    Returns:
        标准行列表（无词片/无翻译的跳过）。
    """
    wb = load_workbook(BytesIO(content), data_only=True)
    if NOTES_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"未找到 sheet「{NOTES_SHEET_NAME}」")
    ws = wb[NOTES_SHEET_NAME]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col: dict[str, int] = {}
    for i, h in enumerate(headers):
        field = NOTES_HEADER_MAP.get(h)
        if field and field not in col:
            col[field] = i

    if "word" not in col:
        raise ValueError("「通用术语」缺少「中文」列")

    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for r in range(2, (ws.max_row or 1) + 1):
        values = [ws.cell(r, c + 1).value for c in range(len(headers))]

        def cell(field: str) -> Any:
            idx = col.get(field)
            return values[idx] if idx is not None else None

        word = str(cell("word") or "").strip()
        if not word:
            continue
        translate = str(cell("translate") or "").strip()
        if not translate:
            continue
        key = (word, target_lang)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "word": word,
                "translate": translate,
                "target_lang": target_lang,
                "department": None,
                "comment": "",
                "category": str(cell("category") or "").strip() or None,
                "abbr": str(cell("abbr") or "").strip() or None,
                "use_llm": parse_use_llm(cell("use_llm")),
                "usage_notes": str(cell("usage_notes") or "").strip() or None,
                "status": default_status if translate else "0",
            }
        )
    return rows
