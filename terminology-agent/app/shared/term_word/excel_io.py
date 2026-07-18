"""术语字典标准 Excel 读写 — 模板 / 导入 / 导出。"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.word_constants import WORD_STATUS_PENDING, WORD_STATUS_UNTRANSLATED
from app.shared.term_word.excel_columns import (
    HEADER_TO_FIELD,
    HEADERS,
    STATUS_CODE_TO_LABEL,
    STATUS_LABEL_TO_CODE,
    USE_LLM_TIP,
)

_SAMPLE_ROWS: list[dict[str, Any]] = [
    {
        "word": "获取/检索",
        "translate": "v. retrieve\nn. retrieval",
        "target_lang": "英文",
        "department": "",
        "comment": "",
        "category": "",
        "abbr": "",
        "use_llm": True,
        "usage_notes": (
            "v.不要用obtain、get、got、parsing、fetching\n"
            "n.不要用acquisition\n"
            "示例：xxx获取失败 → Failed to retrieve xxx"
        ),
        "status": "3",
    },
    {
        "word": "检索器",
        "translate": "Retriever",
        "target_lang": "英文",
        "department": "",
        "comment": "",
        "category": "数据库",
        "abbr": "",
        "use_llm": False,
        "usage_notes": "信息检索领域专指按索引/路径从存储层提取数据的机制。",
        "status": "3",
    },
    {
        "word": "常驻检索器",
        "translate": "Resident retriever",
        "target_lang": "英文",
        "department": "",
        "comment": "",
        "category": "数据库",
        "abbr": "",
        "use_llm": False,
        "usage_notes": "不要用 Resident searcher。",
        "status": "3",
    },
    {
        "word": "画面",
        "translate": "graph",
        "target_lang": "英文",
        "department": "",
        "comment": "",
        "category": "图形",
        "abbr": "",
        "use_llm": False,
        "usage_notes": "不要用 screen、image（自动成图的图也指画面）。",
        "status": "3",
    },
    {
        "word": "图元",
        "translate": "icon",
        "target_lang": "英文",
        "department": "",
        "comment": "",
        "category": "图形",
        "abbr": "",
        "use_llm": False,
        "usage_notes": "不要用 elements、graphic elements、entity 等。",
        "status": "3",
    },
]


def parse_use_llm(value: Any) -> bool:
    """解析走LLM 单元格为 bool。"""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in {"1", "true", "yes", "y", "是", "真"}


def parse_status(value: Any, *, has_translate: bool) -> str:
    """解析翻译状态；空则按是否有翻译给默认。"""
    if value is None or str(value).strip() == "":
        return WORD_STATUS_PENDING if has_translate else WORD_STATUS_UNTRANSLATED
    raw = str(value).strip()
    return STATUS_LABEL_TO_CODE.get(raw, STATUS_LABEL_TO_CODE.get(raw.lower(), WORD_STATUS_PENDING))


def row_to_cells(row: dict[str, Any]) -> list[Any]:
    """标准行 dict → Excel 单元格列表。"""
    status = str(row.get("status") or "")
    status_label = STATUS_CODE_TO_LABEL.get(status, status or "待审核")
    return [
        row.get("word") or "",
        row.get("translate") or "",
        row.get("target_lang") or "",
        row.get("department") or "",
        row.get("comment") or "",
        row.get("category") or "",
        row.get("abbr") or "",
        bool(row.get("use_llm")),
        row.get("usage_notes") or "",
        status_label,
    ]


def _style_header(ws, use_llm_col: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei", size=11)
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        if c == use_llm_col:
            cell.comment = Comment(USE_LLM_TIP, "术语字典", height=80, width=280)


def build_workbook(rows: list[dict[str, Any]] | None = None, *, with_guide: bool = True) -> Workbook:
    """构建标准术语字典 Workbook。"""
    wb = Workbook()
    if with_guide:
        guide = wb.active
        guide.title = "填写说明"
        guide["A1"] = "术语字典标准导入模板"
        guide["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E79")
        lines = [
            "",
            "正式导入只认「术语字典导入」sheet 表头列名。",
            f"走LLM tip：{USE_LLM_TIP}",
            "翻译状态：未翻译/待审核/审核不通过/已审核 或 0/1/2/3。",
            "重复键 (词片, comment, 翻译类型) 已存在则跳过。",
        ]
        for i, line in enumerate(lines, 2):
            guide.cell(i, 1, line)
        guide.column_dimensions["A"].width = 100
        ws = wb.create_sheet("术语字典导入")
    else:
        ws = wb.active
        ws.title = "术语字典导入"

    use_llm_col = HEADERS.index("走LLM") + 1
    _style_header(ws, use_llm_col)
    font = Font(name="Microsoft YaHei", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    for r, row in enumerate(rows or [], 2):
        for c, val in enumerate(row_to_cells(row), 1):
            cell = ws.cell(r, c, val)
            cell.font = font
            cell.alignment = wrap
    widths = [14, 22, 12, 12, 12, 10, 10, 10, 48, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Workbook → xlsx bytes。"""
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_bytes(*, with_sample: bool = True) -> bytes:
    """下载用模板 bytes。"""
    rows = list(_SAMPLE_ROWS) if with_sample else []
    return workbook_to_bytes(build_workbook(rows, with_guide=True))


def parse_import_rows(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """解析上传 xlsx → 标准行列表 + 错误信息。"""
    errors: list[str] = []
    wb = load_workbook(BytesIO(content), data_only=True)
    if "术语字典导入" in wb.sheetnames:
        ws = wb["术语字典导入"]
    else:
        ws = wb.active

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_index: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in HEADER_TO_FIELD:
            col_index[HEADER_TO_FIELD[h]] = i

    if "word" not in col_index or "target_lang" not in col_index:
        return [], ["缺少必填列：词片 / 翻译类型"]

    rows: list[dict[str, Any]] = []
    for r in range(2, (ws.max_row or 1) + 1):
        values = [ws.cell(r, c + 1).value for c in range(len(headers))]
        if not any(v is not None and str(v).strip() for v in values):
            continue

        def cell(field: str) -> Any:
            idx = col_index.get(field)
            return values[idx] if idx is not None else None

        word = str(cell("word") or "").strip()
        target_lang = str(cell("target_lang") or "").strip()
        if not word or not target_lang:
            errors.append(f"第{r}行：词片与翻译类型必填")
            continue
        translate = str(cell("translate") or "").strip()
        status = parse_status(cell("status"), has_translate=bool(translate))
        rows.append(
            {
                "word": word,
                "translate": translate,
                "target_lang": target_lang,
                "department": (str(cell("department")).strip() if cell("department") else None)
                or None,
                "comment": str(cell("comment") or "").strip(),
                "category": str(cell("category") or "").strip() or None,
                "abbr": str(cell("abbr") or "").strip() or None,
                "use_llm": parse_use_llm(cell("use_llm")),
                "usage_notes": str(cell("usage_notes") or "").strip() or None,
                "status": status,
                "_row": r,
            }
        )
    return rows, errors


def demo_sample_path() -> Path | None:
    """仓库内预览样板路径（若存在）。"""
    root = Path(__file__).resolve().parents[4]
    p = root / "docs" / "demos" / "term-word-dictionary" / "term_word_import_sample.xlsx"
    return p if p.is_file() else None
